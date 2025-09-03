import openpyxl
import pandas as pd
from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.pivot import PivotTable, PivotField
from typing import Callable
import os
import tempfile

from app.utils.logger import logger

class BackgroundAutomationService:
    def __init__(self):
        self.workbook = None
        self.worksheet = None
    
    def execute(self, task: str, progress_callback: Callable[[int, str], None]) -> str:
        """Execute a background automation task using openpyxl"""
        try:
            progress_callback(10, "Parsing task requirements...")
            
            # Parse task and determine actions
            actions = self._parse_task(task)
            
            progress_callback(20, "Creating or opening Excel file...")
            
            # Create or open workbook
            self._setup_workbook()
            
            progress_callback(30, "Executing automation tasks...")
            
            # Execute the parsed actions
            result = self._execute_actions(actions, progress_callback)
            
            progress_callback(90, "Saving Excel file...")
            
            # Save the workbook
            output_path = self._save_workbook()
            
            progress_callback(100, "Task completed successfully")
            
            return f"{result}. File saved to: {output_path}"
            
        except Exception as e:
            logger.error(f"Background automation error: {str(e)}")
            raise e
    
    def _parse_task(self, task: str) -> list:
        """Parse task description into actionable steps"""
        task_lower = task.lower()
        actions = []
        
        # Enhanced task parsing for background mode
        if "pivot" in task_lower:
            actions.append({"type": "create_pivot_table", "description": task})
        elif "chart" in task_lower or "graph" in task_lower:
            chart_type = "bar"
            if "line" in task_lower:
                chart_type = "line"
            elif "pie" in task_lower:
                chart_type = "pie"
            actions.append({"type": "create_chart", "chart_type": chart_type, "description": task})
        elif "format" in task_lower:
            actions.append({"type": "format_cells", "description": task})
        elif "formula" in task_lower or "function" in task_lower:
            actions.append({"type": "enter_formulas", "description": task})
        elif "data" in task_lower:
            actions.append({"type": "process_data", "description": task})
        else:
            # Create sample data and perform basic operations
            actions.append({"type": "create_sample_data", "description": task})
        
        return actions
    
    def _setup_workbook(self):
        """Create or open Excel workbook"""
        # Create a new workbook
        self.workbook = openpyxl.Workbook()
        self.worksheet = self.workbook.active
        self.worksheet.title = "Data"
        
        # Add some sample data for demonstration
        self._add_sample_data()
    
    def _add_sample_data(self):
        """Add sample data to the worksheet"""
        # Sample sales data
        headers = ["Product", "Region", "Sales", "Quantity", "Date"]
        data = [
            ["Laptop", "North", 1200, 5, "2024-01-15"],
            ["Mouse", "South", 25, 50, "2024-01-16"],
            ["Keyboard", "East", 75, 20, "2024-01-17"],
            ["Monitor", "West", 300, 8, "2024-01-18"],
            ["Laptop", "South", 1200, 3, "2024-01-19"],
            ["Mouse", "North", 25, 30, "2024-01-20"],
            ["Keyboard", "West", 75, 15, "2024-01-21"],
            ["Monitor", "East", 300, 6, "2024-01-22"],
            ["Laptop", "East", 1200, 7, "2024-01-23"],
            ["Mouse", "West", 25, 40, "2024-01-24"]
        ]
        
        # Write headers
        for col, header in enumerate(headers, 1):
            self.worksheet.cell(row=1, column=col, value=header)
        
        # Write data
        for row, record in enumerate(data, 2):
            for col, value in enumerate(record, 1):
                self.worksheet.cell(row=row, column=col, value=value)
    
    def _execute_actions(self, actions: list, progress_callback: Callable[[int, str], None]) -> str:
        """Execute the parsed actions"""
        results = []
        total_actions = len(actions)
        
        for i, action in enumerate(actions):
            progress = 30 + int((i / total_actions) * 50)  # 30% to 80%
            progress_callback(progress, f"Executing: {action['type']}")
            
            try:
                if action["type"] == "create_pivot_table":
                    result = self._create_pivot_table(action)
                elif action["type"] == "create_chart":
                    result = self._create_chart(action)
                elif action["type"] == "format_cells":
                    result = self._format_cells(action)
                elif action["type"] == "enter_formulas":
                    result = self._enter_formulas(action)
                elif action["type"] == "process_data":
                    result = self._process_data(action)
                elif action["type"] == "create_sample_data":
                    result = self._create_sample_worksheet(action)
                else:
                    result = f"Action {action['type']} completed"
                
                results.append(result)
                
            except Exception as e:
                error_msg = f"Error executing {action['type']}: {str(e)}"
                logger.error(error_msg)
                results.append(error_msg)
        
        return "; ".join(results)
    
    def _create_pivot_table(self, action: dict) -> str:
        """Create a pivot table (simplified implementation)"""
        try:
            # Create a new worksheet for pivot table
            pivot_ws = self.workbook.create_sheet("Pivot Analysis")
            
            # Create a summary table manually (openpyxl doesn't have full pivot table support)
            pivot_ws["A1"] = "Region Summary"
            pivot_ws["A2"] = "Region"
            pivot_ws["B2"] = "Total Sales"
            pivot_ws["C2"] = "Total Quantity"
            
            # Calculate summary data
            regions = {}
            for row in range(2, self.worksheet.max_row + 1):
                region = self.worksheet.cell(row=row, column=2).value
                sales = self.worksheet.cell(row=row, column=3).value
                quantity = self.worksheet.cell(row=row, column=4).value
                
                if region not in regions:
                    regions[region] = {"sales": 0, "quantity": 0}
                regions[region]["sales"] += sales
                regions[region]["quantity"] += quantity
            
            # Write summary data
            current_row = 3
            for region, data in regions.items():
                pivot_ws.cell(row=current_row, column=1, value=region)
                pivot_ws.cell(row=current_row, column=2, value=data["sales"])
                pivot_ws.cell(row=current_row, column=3, value=data["quantity"])
                current_row += 1
            
            # Format the pivot table
            self._format_pivot_table(pivot_ws, current_row - 1)
            
            return "Pivot table analysis created successfully"
            
        except Exception as e:
            raise Exception(f"Failed to create pivot table: {str(e)}")
    
    def _create_chart(self, action: dict) -> str:
        """Create a chart"""
        try:
            chart_type = action.get("chart_type", "bar")
            
            # Create chart based on type
            if chart_type == "line":
                chart = LineChart()
                chart.title = "Sales Trend"
            elif chart_type == "pie":
                chart = PieChart()
                chart.title = "Sales Distribution"
            else:  # default to bar
                chart = BarChart()
                chart.title = "Sales by Region"
            
            # Add data to chart
            data_range = Reference(self.worksheet, min_col=2, min_row=1, max_row=self.worksheet.max_row, max_col=3)
            categories = Reference(self.worksheet, min_col=1, min_row=2, max_row=self.worksheet.max_row)
            
            chart.add_data(data_range, titles_from_data=True)
            if chart_type != "pie":  # Pie charts don't use categories the same way
                chart.set_categories(categories)
            
            # Add chart to worksheet
            chart_ws = self.workbook.create_sheet("Chart")
            chart_ws.add_chart(chart, "A1")
            
            return f"{chart_type.title()} chart created successfully"
            
        except Exception as e:
            raise Exception(f"Failed to create chart: {str(e)}")
    
    def _format_cells(self, action: dict) -> str:
        """Format cells based on task description"""
        try:
            # Apply formatting to headers
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            
            for col in range(1, 6):  # A to E
                cell = self.worksheet.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
            
            # Apply borders to data area
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for row in range(1, self.worksheet.max_row + 1):
                for col in range(1, 6):
                    self.worksheet.cell(row=row, column=col).border = thin_border
            
            # Format sales column as currency
            for row in range(2, self.worksheet.max_row + 1):
                sales_cell = self.worksheet.cell(row=row, column=3)
                sales_cell.number_format = '$#,##0.00'
            
            return "Cell formatting applied successfully"
            
        except Exception as e:
            raise Exception(f"Failed to format cells: {str(e)}")
    
    def _enter_formulas(self, action: dict) -> str:
        """Enter formulas in cells"""
        try:
            # Add summary formulas at the bottom
            last_row = self.worksheet.max_row + 2
            
            self.worksheet.cell(row=last_row, column=1, value="Total Sales:")
            self.worksheet.cell(row=last_row, column=2, value=f"=SUM(C2:C{self.worksheet.max_row-1})")
            
            self.worksheet.cell(row=last_row + 1, column=1, value="Average Sales:")
            self.worksheet.cell(row=last_row + 1, column=2, value=f"=AVERAGE(C2:C{self.worksheet.max_row-2})")
            
            self.worksheet.cell(row=last_row + 2, column=1, value="Total Quantity:")
            self.worksheet.cell(row=last_row + 2, column=2, value=f"=SUM(D2:D{self.worksheet.max_row-3})")
            
            # Format the summary section
            for row in range(last_row, last_row + 3):
                self.worksheet.cell(row=row, column=1).font = Font(bold=True)
                self.worksheet.cell(row=row, column=2).number_format = '#,##0.00'
            
            return "Formulas added successfully"
            
        except Exception as e:
            raise Exception(f"Failed to enter formulas: {str(e)}")
    
    def _process_data(self, action: dict) -> str:
        """Process and manipulate data"""
        try:
            # Add calculated columns
            calc_col = self.worksheet.max_column + 1
            self.worksheet.cell(row=1, column=calc_col, value="Revenue Per Unit")
            
            for row in range(2, self.worksheet.max_row + 1):
                sales = self.worksheet.cell(row=row, column=3).value
                quantity = self.worksheet.cell(row=row, column=4).value
                revenue_per_unit = sales / quantity if quantity > 0 else 0
                self.worksheet.cell(row=row, column=calc_col, value=revenue_per_unit)
            
            return "Data processing completed with calculated columns"
            
        except Exception as e:
            raise Exception(f"Failed to process data: {str(e)}")
    
    def _create_sample_worksheet(self, action: dict) -> str:
        """Create additional sample data worksheet"""
        try:
            # Create new worksheet
            sample_ws = self.workbook.create_sheet("Sample Analysis")
            
            # Add different sample data
            sample_ws["A1"] = "Month"
            sample_ws["B1"] = "Revenue"
            sample_ws["C1"] = "Expenses"
            sample_ws["D1"] = "Profit"
            
            months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun"]
            for i, month in enumerate(months, 2):
                sample_ws.cell(row=i, column=1, value=month)
                revenue = 10000 + (i * 500)
                expenses = 6000 + (i * 300)
                sample_ws.cell(row=i, column=2, value=revenue)
                sample_ws.cell(row=i, column=3, value=expenses)
                sample_ws.cell(row=i, column=4, value=f"=B{i}-C{i}")
            
            return "Sample analysis worksheet created"
            
        except Exception as e:
            raise Exception(f"Failed to create sample worksheet: {str(e)}")
    
    def _format_pivot_table(self, ws, max_row):
        """Format the pivot table"""
        # Format headers
        for col in range(1, 4):
            cell = ws.cell(row=2, column=col)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
        
        # Format title
        title_cell = ws.cell(row=1, column=1)
        title_cell.font = Font(bold=True, size=14)
        
        # Format data area
        for row in range(3, max_row + 1):
            ws.cell(row=row, column=2).number_format = '$#,##0'
            ws.cell(row=row, column=3).number_format = '#,##0'
    
    def _save_workbook(self) -> str:
        """Save the workbook to a file"""
        try:
            # Create output directory if it doesn't exist
            output_dir = "output"
            if not os.path.exists(output_dir):
                os.makedirs(output_dir)
            
            # Generate filename with timestamp
            import time
            timestamp = int(time.time())
            filename = f"nubia_automation_{timestamp}.xlsx"
            filepath = os.path.join(output_dir, filename)
            
            # Save workbook
            self.workbook.save(filepath)
            
            return os.path.abspath(filepath)
            
        except Exception as e:
            # Fallback to temp directory
            temp_dir = tempfile.gettempdir()
            timestamp = int(time.time())
            filename = f"nubia_automation_{timestamp}.xlsx"
            filepath = os.path.join(temp_dir, filename)
            
            self.workbook.save(filepath)
            
            return os.path.abspath(filepath)