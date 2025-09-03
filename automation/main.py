#!/usr/bin/env python3

import asyncio
import json
import logging
import os
import sys
import time
import traceback
from datetime import datetime
from typing import Dict, List, Optional, Any
import uuid

import pyautogui
from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, LineChart, PieChart, ScatterChart, Reference
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
import pandas as pd

from fastapi import FastAPI, HTTPException, WebSocket, WebSocketDisconnect
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
import uvicorn

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Initialize FastAPI
app = FastAPI(title="Nubia Excel Automation Service", version="1.0.0")

# Add CORS middleware
app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:3000", "http://127.0.0.1:3000"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Global state
active_tasks: Dict[str, "AutomationTask"] = {}
connected_websockets: List[WebSocket] = []

# Pydantic models
class ExcelStep(BaseModel):
    id: str
    action: str
    description: str
    parameters: Dict[str, Any]
    order: int
    estimatedTime: int

class ExcelTask(BaseModel):
    id: str
    type: str
    description: str
    complexity: str
    estimatedActions: int
    parameters: Dict[str, Any]
    mode: str  # 'visual' or 'background'
    priority: int
    status: str
    steps: List[ExcelStep]
    metadata: Dict[str, Any]

class ExecuteRequest(BaseModel):
    task: ExcelTask

class TaskStatus(BaseModel):
    task_id: str
    status: str
    progress: int
    current_step: Optional[str] = None
    message: Optional[str] = None
    error: Optional[str] = None

# Automation task execution class
class AutomationTask:
    def __init__(self, task: ExcelTask):
        self.task = task
        self.status = "pending"
        self.progress = 0
        self.current_step = None
        self.error = None
        self.workbook = None
        self.worksheet = None
        self.start_time = None
        self.end_time = None

    async def execute(self):
        """Execute the automation task"""
        try:
            self.status = "in_progress"
            self.start_time = datetime.now()
            logger.info(f"Starting execution of task {self.task.id}")

            if self.task.mode == "visual":
                await self.execute_visual_mode()
            else:
                await self.execute_background_mode()

            self.status = "completed"
            self.progress = 100
            self.end_time = datetime.now()
            logger.info(f"Task {self.task.id} completed successfully")

        except Exception as e:
            self.status = "failed"
            self.error = str(e)
            logger.error(f"Task {self.task.id} failed: {e}")
            logger.error(traceback.format_exc())
        finally:
            await self.broadcast_status()

    async def execute_visual_mode(self):
        """Execute task using pyautogui for visual automation"""
        logger.info("Executing in visual mode using pyautogui")
        
        # Configure pyautogui
        pyautogui.FAILSAFE = True
        pyautogui.PAUSE = 0.5  # Pause between actions
        
        try:
            # Check if Excel is available
            if not self.check_excel_available():
                raise Exception("Microsoft Excel is not installed or not accessible")

            for i, step in enumerate(self.task.steps):
                self.current_step = step.description
                await self.broadcast_status()
                
                await self.execute_visual_step(step)
                
                # Update progress
                self.progress = int((i + 1) / len(self.task.steps) * 100)
                await self.broadcast_status()
                
                # Small delay between steps
                await asyncio.sleep(0.5)

        except Exception as e:
            logger.error(f"Visual mode execution failed: {e}")
            raise

    async def execute_background_mode(self):
        """Execute task using openpyxl for background automation"""
        logger.info("Executing in background mode using openpyxl")
        
        try:
            for i, step in enumerate(self.task.steps):
                self.current_step = step.description
                await self.broadcast_status()
                
                await self.execute_background_step(step)
                
                # Update progress
                self.progress = int((i + 1) / len(self.task.steps) * 100)
                await self.broadcast_status()

        except Exception as e:
            logger.error(f"Background mode execution failed: {e}")
            raise

    async def execute_visual_step(self, step: ExcelStep):
        """Execute a single step in visual mode"""
        action = step.action
        params = step.parameters

        logger.info(f"Executing visual step: {action}")

        if action == "new_workbook":
            await self.visual_new_workbook()
        elif action == "add_worksheet":
            await self.visual_add_worksheet(params.get("name", "Sheet1"))
        elif action == "enter_data":
            await self.visual_enter_data(params)
        elif action == "create_chart":
            await self.visual_create_chart(params)
        elif action == "save_file":
            await self.visual_save_file(params.get("fileName"))
        else:
            logger.warning(f"Unknown visual action: {action}")

    async def execute_background_step(self, step: ExcelStep):
        """Execute a single step in background mode"""
        action = step.action
        params = step.parameters

        logger.info(f"Executing background step: {action}")

        if action == "new_workbook":
            self.workbook = Workbook()
            self.worksheet = self.workbook.active
        elif action == "add_worksheet":
            if self.workbook:
                self.worksheet = self.workbook.create_sheet(params.get("name", "Sheet1"))
        elif action == "enter_data":
            await self.background_enter_data(params)
        elif action == "create_formula":
            await self.background_create_formula(params)
        elif action == "create_chart":
            await self.background_create_chart(params)
        elif action == "format_range":
            await self.background_format_range(params)
        elif action == "save_file":
            await self.background_save_file(params.get("fileName"))
        else:
            logger.warning(f"Unknown background action: {action}")

    # Visual mode implementations
    async def visual_new_workbook(self):
        """Open new Excel workbook visually"""
        try:
            # Try to open Excel (this is platform-specific)
            if sys.platform == "win32":
                pyautogui.press('win')
                pyautogui.write('excel')
                pyautogui.press('enter')
                await asyncio.sleep(3)
                pyautogui.hotkey('ctrl', 'n')
            elif sys.platform == "darwin":  # macOS
                pyautogui.hotkey('cmd', 'space')
                pyautogui.write('excel')
                pyautogui.press('enter')
                await asyncio.sleep(3)
                pyautogui.hotkey('cmd', 'n')
            else:
                raise Exception("Visual mode not supported on this platform")
                
        except Exception as e:
            logger.warning(f"Failed to open Excel visually: {e}")
            raise Exception("Could not open Excel application")

    async def visual_add_worksheet(self, name: str):
        """Add worksheet in Excel visually"""
        try:
            # Right-click on sheet tab and insert new worksheet
            pyautogui.hotkey('shift', 'f11')  # Insert worksheet shortcut
            await asyncio.sleep(1)
        except Exception as e:
            logger.warning(f"Failed to add worksheet: {e}")

    async def visual_enter_data(self, params: Dict[str, Any]):
        """Enter data into Excel cells visually"""
        try:
            data = params.get("data", [])
            range_start = params.get("range", "A1")
            
            # Click on starting cell
            if range_start:
                pyautogui.hotkey('ctrl', 'g')  # Go to dialog
                await asyncio.sleep(0.5)
                pyautogui.write(range_start)
                pyautogui.press('enter')
                await asyncio.sleep(0.5)

            # Enter data row by row
            for row_idx, row in enumerate(data):
                for col_idx, cell_value in enumerate(row):
                    if col_idx > 0:
                        pyautogui.press('tab')
                    pyautogui.write(str(cell_value))
                    
                if row_idx < len(data) - 1:
                    pyautogui.press('enter')
                    # Move back to first column
                    for _ in range(len(row) - 1):
                        pyautogui.press('left')

        except Exception as e:
            logger.warning(f"Failed to enter data: {e}")

    async def visual_create_chart(self, params: Dict[str, Any]):
        """Create chart in Excel visually"""
        try:
            chart_type = params.get("type", "column")
            
            # Select data first (assuming it's already selected)
            pyautogui.hotkey('alt', 'n', 'c')  # Insert chart
            await asyncio.sleep(1)
            
            # Choose chart type (this varies by Excel version)
            if chart_type == "pie":
                pyautogui.press('p')
            elif chart_type == "line":
                pyautogui.press('l')
            else:  # default to column
                pyautogui.press('c')
                
            pyautogui.press('enter')
            await asyncio.sleep(2)
            
        except Exception as e:
            logger.warning(f"Failed to create chart: {e}")

    async def visual_save_file(self, filename: str):
        """Save file in Excel visually"""
        try:
            pyautogui.hotkey('ctrl', 's')
            await asyncio.sleep(1)
            
            if filename:
                pyautogui.write(filename)
                
            pyautogui.press('enter')
            await asyncio.sleep(1)
            
        except Exception as e:
            logger.warning(f"Failed to save file: {e}")

    # Background mode implementations
    async def background_enter_data(self, params: Dict[str, Any]):
        """Enter data using openpyxl"""
        if not self.worksheet:
            return

        data = params.get("data", [])
        range_start = params.get("range", "A1")
        
        # Parse starting position
        col_start = 1
        row_start = 1
        if range_start:
            # Simple parsing for A1 notation
            import re
            match = re.match(r'([A-Z]+)(\d+)', range_start)
            if match:
                col_letters = match.group(1)
                row_start = int(match.group(2))
                col_start = sum((ord(c) - ord('A') + 1) * (26 ** i) for i, c in enumerate(reversed(col_letters.upper())))

        # Enter data
        for row_idx, row in enumerate(data):
            for col_idx, cell_value in enumerate(row):
                cell = self.worksheet.cell(
                    row=row_start + row_idx,
                    column=col_start + col_idx,
                    value=cell_value
                )

    async def background_create_formula(self, params: Dict[str, Any]):
        """Create formula using openpyxl"""
        if not self.worksheet:
            return

        formula = params.get("formula", "=SUM(A1:A10)")
        cell_range = params.get("range", "B1")
        
        # Parse cell position
        import re
        match = re.match(r'([A-Z]+)(\d+)', cell_range)
        if match:
            col_letters = match.group(1)
            row = int(match.group(2))
            col = sum((ord(c) - ord('A') + 1) * (26 ** i) for i, c in enumerate(reversed(col_letters.upper())))
            
            self.worksheet.cell(row=row, column=col, value=formula)

    async def background_create_chart(self, params: Dict[str, Any]):
        """Create chart using openpyxl"""
        if not self.worksheet:
            return

        chart_type = params.get("type", "column")
        data_range = params.get("range", "A1:B10")
        
        # Create appropriate chart
        if chart_type == "bar":
            chart = BarChart()
        elif chart_type == "line":
            chart = LineChart()
        elif chart_type == "pie":
            chart = PieChart()
        elif chart_type == "scatter":
            chart = ScatterChart()
        else:
            chart = BarChart()  # default

        # Add data to chart
        data = Reference(self.worksheet, range_string=data_range)
        chart.add_data(data, titles_from_data=True)
        
        # Add chart to worksheet
        self.worksheet.add_chart(chart, "E5")

    async def background_format_range(self, params: Dict[str, Any]):
        """Format cells using openpyxl"""
        if not self.worksheet:
            return

        cell_range = params.get("range", "A1")
        formatting = params.get("formatting", {})
        
        # Apply formatting
        for row in self.worksheet[cell_range]:
            for cell in (row if isinstance(row, tuple) else [row]):
                if formatting.get("bold"):
                    cell.font = Font(bold=True)
                if formatting.get("backgroundColor"):
                    cell.fill = PatternFill(start_color=formatting["backgroundColor"], 
                                          end_color=formatting["backgroundColor"], 
                                          fill_type="solid")

    async def background_save_file(self, filename: str):
        """Save workbook using openpyxl"""
        if not self.workbook or not filename:
            return

        try:
            # Ensure filename has .xlsx extension
            if not filename.endswith('.xlsx'):
                filename += '.xlsx'
            
            # Create output directory if it doesn't exist
            output_dir = os.path.join(os.path.dirname(__file__), "output")
            os.makedirs(output_dir, exist_ok=True)
            
            filepath = os.path.join(output_dir, filename)
            self.workbook.save(filepath)
            logger.info(f"File saved to {filepath}")
            
        except Exception as e:
            logger.error(f"Failed to save file: {e}")
            raise

    def check_excel_available(self) -> bool:
        """Check if Excel is available on the system"""
        try:
            if sys.platform == "win32":
                import winreg
                try:
                    key = winreg.OpenKey(winreg.HKEY_LOCAL_MACHINE, 
                                        r"SOFTWARE\Microsoft\Office")
                    winreg.CloseKey(key)
                    return True
                except:
                    return False
            elif sys.platform == "darwin":
                # Check for Excel on macOS
                return os.path.exists("/Applications/Microsoft Excel.app")
            else:
                # Linux - check for LibreOffice Calc
                import shutil
                return shutil.which("libreoffice") is not None
        except:
            return False

    async def broadcast_status(self):
        """Broadcast task status to connected WebSocket clients"""
        status = TaskStatus(
            task_id=self.task.id,
            status=self.status,
            progress=self.progress,
            current_step=self.current_step,
            error=self.error
        )
        
        message = status.model_dump_json()
        
        # Send to all connected WebSocket clients
        disconnected = []
        for websocket in connected_websockets:
            try:
                await websocket.send_text(message)
            except:
                disconnected.append(websocket)
        
        # Remove disconnected clients
        for ws in disconnected:
            connected_websockets.remove(ws)

# API Endpoints
@app.get("/")
async def root():
    return {"message": "Nubia Excel Automation Service", "version": "1.0.0", "status": "running"}

@app.get("/health")
async def health():
    return {
        "status": "healthy",
        "timestamp": datetime.now().isoformat(),
        "active_tasks": len(active_tasks),
        "excel_available": check_excel_system()
    }

@app.post("/execute")
async def execute_task(request: ExecuteRequest):
    """Execute an Excel automation task"""
    try:
        task = request.task
        
        if task.id in active_tasks:
            raise HTTPException(status_code=400, detail="Task already running")
        
        # Create and start automation task
        automation_task = AutomationTask(task)
        active_tasks[task.id] = automation_task
        
        # Execute in background
        asyncio.create_task(automation_task.execute())
        
        return {
            "message": "Task started",
            "task_id": task.id,
            "mode": task.mode,
            "estimated_duration": task.metadata.get("estimatedDuration", 60)
        }
        
    except Exception as e:
        logger.error(f"Failed to execute task: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/status/{task_id}")
async def get_task_status(task_id: str):
    """Get the status of a specific task"""
    if task_id not in active_tasks:
        raise HTTPException(status_code=404, detail="Task not found")
    
    task = active_tasks[task_id]
    return TaskStatus(
        task_id=task_id,
        status=task.status,
        progress=task.progress,
        current_step=task.current_step,
        error=task.error
    )

@app.post("/abort/{task_id}")
async def abort_task(task_id: str):
    """Abort a running task"""
    if task_id not in active_tasks:
        raise HTTPException(status_code=404, detail="Task not found")
    
    task = active_tasks[task_id]
    task.status = "aborted"
    task.error = "Task aborted by user"
    
    # Remove from active tasks
    del active_tasks[task_id]
    
    return {"message": "Task aborted", "task_id": task_id}

@app.get("/tasks")
async def list_active_tasks():
    """List all active tasks"""
    tasks = []
    for task_id, task in active_tasks.items():
        tasks.append({
            "task_id": task_id,
            "description": task.task.description,
            "status": task.status,
            "progress": task.progress,
            "mode": task.task.mode,
            "start_time": task.start_time.isoformat() if task.start_time else None
        })
    return {"active_tasks": tasks}

@app.websocket("/ws")
async def websocket_endpoint(websocket: WebSocket):
    """WebSocket endpoint for real-time task updates"""
    await websocket.accept()
    connected_websockets.append(websocket)
    
    try:
        while True:
            # Keep connection alive
            await websocket.receive_text()
    except WebSocketDisconnect:
        connected_websockets.remove(websocket)

def check_excel_system() -> bool:
    """Check if Excel is available on the system"""
    try:
        task = AutomationTask(ExcelTask(
            id="temp", type="check", description="", complexity="simple",
            estimatedActions=1, parameters={}, mode="visual", priority=1,
            status="pending", steps=[], metadata={}
        ))
        return task.check_excel_available()
    except:
        return False

if __name__ == "__main__":
    # Create output directory
    os.makedirs("output", exist_ok=True)
    
    # Check system requirements
    logger.info("Starting Nubia Excel Automation Service...")
    logger.info(f"Excel available: {check_excel_system()}")
    logger.info(f"Platform: {sys.platform}")
    
    # Run the server
    uvicorn.run(
        "main:app",
        host="127.0.0.1",
        port=8000,
        reload=False,
        log_level="info"
    )